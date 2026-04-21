# -*- coding: utf-8 -*-
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

xlsx_file = 'Family_Event_Coordination_Template_v0.2.xlsx'
wb = load_workbook(xlsx_file)

# Remove schema sheet if it exists
if 'Schema Design' in wb.sheetnames:
    del wb['Schema Design']

# Create new schema sheet
schema_sheet = wb.create_sheet('Schema Design', 1)

# Define styles
header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
subheader_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
subheader_font = Font(bold=True, color='FFFFFF', size=10)
entity_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
entity_font = Font(bold=True, size=10)

# Set column widths
schema_sheet.column_dimensions['A'].width = 25
schema_sheet.column_dimensions['B'].width = 50
schema_sheet.column_dimensions['C'].width = 15

row = 1

# Title
title_cell = schema_sheet[f'A{row}']
title_cell.value = 'SCHEMA DESIGN - Family Event Coordination System'
title_cell.font = Font(bold=True, size=14, color='FFFFFF')
title_cell.fill = header_fill
schema_sheet.merge_cells(f'A{row}:C{row}')
row += 2

# Define all entities
entities = [
    {
        'section': 'CORE: People Management',
        'tables': [
            {
                'name': 'Users',
                'description': 'Account holders. Phone-first identity.',
                'columns': ['user_id (PK)', 'phone_number (UNIQUE)', 'email', 'display_name', 'created_at', 'last_active_at', 'notes'],
                'relationships': '1:N → Families (as owner); N:M → Sides via SideMembership'
            },
            {
                'name': 'Families',
                'description': 'Family unit. Owns members; one Family = one Side.',
                'columns': ['family_id (PK)', 'family_name', 'owner_user_id (FK→Users)', 'home_city', 'created_at', 'notes'],
                'relationships': '1:N → FamilyMembers'
            },
            {
                'name': 'FamilyMembers',
                'description': 'Individual members of a family.',
                'columns': ['member_id (PK)', 'family_id (FK→Families)', 'user_id (FK→Users, optional)', 'full_name', 'phone_number', 'relationship_to_owner', 'is_decision_maker', 'notes'],
                'relationships': 'N:M → Relationships (bidirectional)'
            },
            {
                'name': 'Relationships',
                'description': 'Member-to-member edges (mother, father, sister, etc.).',
                'columns': ['relationship_id (PK)', 'from_member_id (FK→FamilyMembers)', 'to_member_id (FK→FamilyMembers)', 'relationship_type', 'notes'],
                'relationships': 'Self-referencing via FamilyMembers'
            }
        ]
    },
    {
        'section': 'WEDDING & EVENTS: Core Structure',
        'tables': [
            {
                'name': 'Wedding',
                'description': 'Top-level container for the entire event.',
                'columns': ['wedding_id (PK)', 'wedding_name', 'start_date', 'end_date', 'primary_city', 'status', 'is_two_sided', 'notes'],
                'relationships': '1:N → Sides; 1:N → SubEvents'
            },
            {
                'name': 'Sides',
                'description': 'Links one Family to one Wedding (Bride/Groom/Solo).',
                'columns': ['side_id (PK)', 'wedding_id (FK→Wedding)', 'family_id (FK→Families)', 'side_name', 'side_type'],
                'relationships': 'N:M → Users via SideMembership'
            },
            {
                'name': 'SideMembership',
                'description': 'User × Side × Role. Drives all permissions.',
                'columns': ['membership_id (PK)', 'user_id (FK→Users)', 'side_id (FK→Sides)', 'role', 'joined_at'],
                'relationships': 'Permission control for all side-specific data'
            },
            {
                'name': 'SubEvents',
                'description': 'Functions like Mehendi, Sangeet, Ceremony, etc.',
                'columns': ['sub_event_id (PK)', 'wedding_id (FK→Wedding)', 'host_side_id (FK→Sides)', 'event_type', 'event_date', 'start_time', 'end_time', 'location', 'hosting_scope', 'lead_member_id (FK→FamilyMembers)', 'notes'],
                'relationships': '1:N → Bookings; 1:N → GuestInvitations; 1:N → SubEventRequirements; 1:N → Performances'
            }
        ]
    },
    {
        'section': 'VENDOR & FINANCE: Bookings & Costs',
        'tables': [
            {
                'name': 'VendorCategories',
                'description': 'Pre-seeded categories (Catering, Photography, etc.).',
                'columns': ['category_id (PK)', 'category_name', 'parent_category_id (FK→VendorCategories, optional)'],
                'relationships': '1:N → Vendors (lookup reference)'
            },
            {
                'name': 'Vendors',
                'description': 'Vendor contact records. Owned by one Side.',
                'columns': ['vendor_id (PK)', 'vendor_name', 'category_id (FK→VendorCategories)', 'contact_person', 'phone', 'email', 'address', 'city', 'created_by_side_id (FK→Sides)', 'created_at', 'notes'],
                'relationships': '1:N → Bookings'
            },
            {
                'name': 'Bookings',
                'description': 'Vendor × SubEvent. Tracks amount, advance, status.',
                'columns': ['booking_id (PK)', 'vendor_id (FK→Vendors)', 'sub_event_id (FK→SubEvents)', 'booking_date', 'confirmed_amount', 'advance_paid', 'payment_status', 'booking_status', 'notes'],
                'relationships': 'Many-to-many vendor-to-event assignments'
            }
        ]
    },
    {
        'section': 'GUESTS: Invitations & RSVPs',
        'tables': [
            {
                'name': 'Guests',
                'description': 'Invited people. Separate from FamilyMembers.',
                'columns': ['guest_id (PK)', 'full_name', 'phone_number', 'email', 'invited_by_side_id (FK→Sides)', 'relationship', 'dietary_restrictions', 'accessibility_needs', 'created_at', 'notes'],
                'relationships': '1:N → GuestInvitations'
            },
            {
                'name': 'GuestInvitations',
                'description': 'Which guests are invited to which sub-events.',
                'columns': ['invitation_id (PK)', 'guest_id (FK→Guests)', 'sub_event_id (FK→SubEvents)', 'status (pending/accepted/declined)', 'dietary_confirmed', 'plus_ones', 'created_at', 'notes'],
                'relationships': 'Many-to-many guest-to-event invitations'
            }
        ]
    },
    {
        'section': 'ACCOMMODATION: Properties & Rooms',
        'tables': [
            {
                'name': 'Properties',
                'description': 'Hotels, family homes, etc. used for guest stays.',
                'columns': ['property_id (PK)', 'wedding_id (FK→Wedding)', 'property_name', 'address', 'city', 'managed_by_side_id (FK→Sides)', 'check_in_date', 'check_out_date', 'capacity', 'notes'],
                'relationships': '1:N → Rooms'
            },
            {
                'name': 'Rooms',
                'description': 'Individual rooms within a Property.',
                'columns': ['room_id (PK)', 'property_id (FK→Properties)', 'room_number', 'room_type', 'capacity', 'amenities', 'notes'],
                'relationships': '1:N → RoomAssignments'
            },
            {
                'name': 'RoomAssignments',
                'description': 'Guest → Room with check-in/out dates.',
                'columns': ['assignment_id (PK)', 'room_id (FK→Rooms)', 'guest_id (FK→Guests)', 'check_in', 'check_out', 'notes'],
                'relationships': 'Many-to-many guest-to-room assignments'
            }
        ]
    },
    {
        'section': 'ENTERTAINMENT: Performances & Tasks',
        'tables': [
            {
                'name': 'RequirementTemplates',
                'description': 'Pre-seeded ritual checklists per sub-event type.',
                'columns': ['template_id (PK)', 'event_type', 'requirement_description', 'priority'],
                'relationships': '1:N → SubEventRequirements (reference)'
            },
            {
                'name': 'SubEventRequirements',
                'description': 'Instantiated checklist per sub-event.',
                'columns': ['requirement_id (PK)', 'sub_event_id (FK→SubEvents)', 'created_from_template_id (FK→RequirementTemplates)', 'description', 'responsible_member_id (FK→FamilyMembers)', 'status (pending/in_progress/completed)', 'due_date', 'notes'],
                'relationships': 'Task management per sub-event'
            },
            {
                'name': 'Performances',
                'description': 'Performance items (dances, speeches, etc.) for Sangeet.',
                'columns': ['performance_id (PK)', 'sub_event_id (FK→SubEvents)', 'performance_name', 'performance_type', 'planned_duration', 'notes'],
                'relationships': '1:N → Performers'
            },
            {
                'name': 'Performers',
                'description': 'Members or guests participating in a performance.',
                'columns': ['performer_id (PK)', 'performance_id (FK→Performances)', 'member_id (FK→FamilyMembers, optional)', 'guest_id (FK→Guests, optional)', 'role', 'costume_notes', 'notes'],
                'relationships': 'Many-to-many member/guest-to-performance'
            }
        ]
    },
    {
        'section': 'FINANCE & GIFTS: Gift Tracking',
        'tables': [
            {
                'name': 'Gifts',
                'description': 'Shagun envelopes, cash, jewellery received. STRICTLY private.',
                'columns': ['gift_id (PK)', 'wedding_id (FK→Wedding)', 'recorded_by_side_id (FK→Sides)', 'linked_guest_id (FK→Guests, optional)', 'gift_type (cash/jewelry/item)', 'amount_or_description', 'date_received', 'acknowledgment_status', 'notes'],
                'relationships': 'Privacy: OWN SIDE ONLY'
            }
        ]
    }
]

# Add entities to sheet
for entity_section in entities:
    # Section header
    section_cell = schema_sheet[f'A{row}']
    section_cell.value = entity_section['section']
    section_cell.font = subheader_font
    section_cell.fill = subheader_fill
    schema_sheet.merge_cells(f'A{row}:C{row}')
    row += 1

    for table in entity_section['tables']:
        # Table name
        name_cell = schema_sheet[f'A{row}']
        name_cell.value = table['name']
        name_cell.font = entity_font
        name_cell.fill = entity_fill
        schema_sheet.merge_cells(f'A{row}:C{row}')
        row += 1

        # Description
        schema_sheet[f'A{row}'].value = 'Description:'
        schema_sheet[f'B{row}'].value = table['description']
        schema_sheet[f'B{row}'].alignment = Alignment(wrap_text=True)
        row += 1

        # Columns
        schema_sheet[f'A{row}'].value = 'Columns:'
        row += 1
        for col in table['columns']:
            schema_sheet[f'B{row}'].value = col
            row += 1

        # Relationships
        schema_sheet[f'A{row}'].value = 'Relationships:'
        schema_sheet[f'B{row}'].value = table['relationships']
        schema_sheet[f'B{row}'].alignment = Alignment(wrap_text=True)
        row += 2

wb.save(xlsx_file)
print("[OK] Schema Design sheet created successfully!")
